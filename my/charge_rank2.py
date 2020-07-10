#!/usr/bin/python3
#-*- coding: utf-8 -*-

from math import ceil
import mysql.connector
import os
from datetime import datetime,timedelta
from pprint import pprint
import time
import openpyxl
import xlrd
import xlwt
import sys
import argparse
from xlutils.copy import copy

config = {
      'user': 'root',
      'password': '',
      'host': 'localhost',
      'database': 'cj655',
      'charset': 'utf8',
      "connection_timeout": 60,
      "use_pure": True
}

if os.getenv('OS') == 'Windows_NT':
    config['password'] = ''
else:
    config['password'] = 'game123456'

def get_user_id(data):
    user_ids = []
    for item in data:
        user_ids.append(str(item["user_id"]))
    return user_ids

if __name__ == '__main__':
    sd = datetime(2019,5,1)
    start_time = int(sd.timestamp())
    y = datetime.now() - timedelta(days=1)
    ed = datetime(y.year,y.month,y.day,23,59,59)
    end_time = int(ed.timestamp())

    file_name = 'charge_rank-%s-%s.xlsx' % (sd.strftime("%Y-%m-%d"), ed.strftime("%Y-%m-%d"))
    log_dir_name = "./log"
    log_name = "%s/cr-%s" % (log_dir_name,datetime.now().strftime("%Y-%m-%d"))

    if not os.path.exists(log_dir_name):
        os.mkdir(log_dir_name)

    if os.path.exists(file_name):
        os.remove(file_name)

    log_file = open(log_name, 'w')
    log_file.write("\n===== Begin Log %s =====\n" % datetime.now().strftime("%Y-%m-%d"))

    parser = argparse.ArgumentParser()
    parser.add_argument('--page', '-p', help='执行的页码',default=0)
    parser.add_argument('--limit', '-l', help='每页的条数', default=0)
    parser.add_argument('--offset', '-s', help='页码偏移量', default=0)
    args = parser.parse_args()

    conn = mysql.connector.connect(**config)
    cursor = conn.cursor(dictionary=True)

    querySql = "SELECT count(distinct user_id) total_count " \
               "FROM gc_order " \
               "WHERE channel = 1 AND status = 1 AND create_time BETWEEN %d AND %d" % (start_time,end_time)
    cursor.execute(querySql)
    count = cursor.fetchone()
    total_count = count["total_count"]
    print("Total count : %d" % total_count)
    list_rows = 5000
    if int(args.limit) > 0:
        list_rows = int(args.limit)
    total_page = ceil(total_count / list_rows)

    all_start_time = time.time()

    for page in range(1,total_page+1):
        if (int(args.page) > 0) and (page > int(args.page)):break
        if (int(args.offset) > 0) and (page <= int(args.offset)):continue
        #if page != total_page-1: continue

        first_row = (page - 1) * list_rows

        page_start_time = time.time()

        querySql = "SELECT o.user_id,o.username,round(sum(o.money),2) as charge_sum," \
                   "group_concat(distinct g.game_name) game_name," \
                   "group_concat(distinct o.game_server_name) server_name," \
                   "group_concat(distinct o.role_name) role_name," \
                   "u.game_type " \
                   "FROM gc_order as o " \
                   "LEFT JOIN gc_game as g on g.game_id = o.game_id " \
                   "LEFT JOIN gc_user as u on u.user_id = o.user_id " \
                   "WHERE o.channel = 1 AND o.status = 1 AND o.create_time BETWEEN %d AND %d " \
                   "GROUP BY o.user_id ORDER BY charge_sum DESC LIMIT %d,%d" % (start_time, end_time,first_row,list_rows)

        #print(querySql)
        #sys.exit(0)

        cursor.execute(querySql)
        order = cursor.fetchall()

        user_ids = get_user_id(order)

        querySql = "SELECT max(login_time) login_time,user_id FROM gc_user_play_data WHERE user_id in (%s) GROUP BY user_id" % ",".join(user_ids)
        cursor.execute(querySql)
        user_play_data = cursor.fetchall()

        querySql = "SELECT max(create_time) create_time,user_id FROM gc_order WHERE user_id in (%s) AND status = 1 AND channel = 1 GROUP BY user_id" % ",".join(user_ids)
        cursor.execute(querySql)
        order2 = cursor.fetchall()

        #pprint(order2)
        #sys.exit(0)

        for index, value in enumerate(order):
            # 未登录天数
            user_no_login_time = 0
            for value2 in user_play_data:
                if value2["user_id"] == value["user_id"]:
                    user_no_login_time = value2["login_time"]
                    break
            no_login = int((time.time() - user_no_login_time) / 86400)
            order[index]["no_login"] = no_login

            # 未充值天数
            user_no_charge_time = 0
            for value2 in order2:
                if value2["user_id"] == value["user_id"]:
                    user_no_charge_time = value2["create_time"]
                    break
            no_charge = int((time.time() - user_no_charge_time) / 86400)
            order[index]["no_charge"] = no_charge

            # 设备
            if value["game_type"] == 1:
                order[index]["game_type"] = "android"
            else:
                order[index]["game_type"] = "ios"

        #print(len(order))
        #pprint(order)

        if os.path.exists(file_name):
            workbook = openpyxl.load_workbook(file_name)
            worksheet = workbook.worksheets[0]
            row = worksheet.max_row
            row = row+1
            #print("row : %d" % row)
        else:

            workbook = openpyxl.Workbook()
            worksheet = workbook.active

            worksheet["A1"] = '用户ID'
            worksheet["B1"] = '用户名'
            worksheet["C1"] = '设备'
            worksheet["D1"] = '游戏名称'
            worksheet["E1"] = '区服名称'
            worksheet["F1"] = '角色名称'
            worksheet["G1"] = '金额'
            worksheet["H1"] = '未登录天数'
            worksheet["I1"] = '未充值天数'

            row = 2

        for index, value in enumerate(order):
            worksheet["A"+str(row+index)] = str(value["user_id"])
            worksheet["B"+str(row+index)] = str(value["username"])
            worksheet["C"+str(row+index)] = value["game_type"]
            worksheet["D"+str(row+index)] = value["game_name"]
            worksheet["E"+str(row+index)] = value["server_name"]
            worksheet["F"+str(row+index)] = value["role_name"]
            worksheet["G"+str(row+index)] = str(value["charge_sum"])
            worksheet["H"+str(row+index)] = value["no_login"]
            worksheet["I"+str(row+index)] = value["no_charge"]

        workbook.save(file_name)

        page_end_time = time.time()

        print("Page %d is completed,time : %s" % (page,timedelta(seconds=page_end_time-page_start_time)))
        log_file.write("Page %d is completed,time : %s \n" % (page,timedelta(seconds=page_end_time-page_start_time)))
        log_file.flush()

    all_end_time = time.time()

    log_file.write("All page is completed,time : %s \n" % timedelta(seconds=all_end_time - all_start_time))
    log_file.write("===== End Log %s =====" % datetime.now().strftime("%Y-%m-%d"))
    log_file.flush()

